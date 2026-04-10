from django.shortcuts import render, get_object_or_404, redirect
from .models import Producto
from .models import Cliente
from .models import Pedido
from .models import DetallePedido

from .forms import ClienteForm
from .forms import ProductoForm
from .forms import PedidoForm
from .forms import DetallePedidoForm
# importacion de inlineformset_factory para crear un formset de detalles de pedido dentro del formulario de pedido
from django.forms import inlineformset_factory

import requests

from django.core.paginator import Paginator
from io import BytesIO
from django.http import HttpResponse, JsonResponse
# reportes en pdf
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
# reportes en excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# para el dashboard
from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncMonth


# Create your views here.

API_URL = 'http://localhost:8001'

# funcion para listar clientes

def listar_clientes(request):
    # obtenemos todos los clientes de la base de datos
    clientes = Cliente.objects.all()
    # paginamos los clientes
    paginator = Paginator(clientes, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'clientes/listar_clientes.html', {'page_obj': page_obj})

# funcion para listar productos
def listar_productos(request):
    # obtenemos todos los productos de la base de datos
    productos = Producto.objects.all()
    # paginamos los productos
    paginator = Paginator(productos, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'productos/listar_productos.html', {'page_obj': page_obj})

# funcion para listar pedidos
def listar_pedidos(request):
    # obtenemos todos los pedidos de la base de datos
    pedidos = Pedido.objects.all().order_by('-id')
    # paginamos los pedidos
    paginator = Paginator(pedidos, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'pedidos/listar_pedidos.html', {'page_obj': page_obj})

# funcion para listar detalles de pedidos
def listar_detalles_pedidos(request):
    # obtenemos todos los detalles de pedidos de la base de datos
    detalles_pedidos = DetallePedido.objects.all()
    # paginamos los detalles de pedidos
    paginator = Paginator(detalles_pedidos, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'detalles_pedidos/listar_detalles_pedidos.html', {'page_obj': page_obj})

# funcion para crear un nuevo cliente
def crear_cliente(request):
    # si es get y no es ajax, redirigimos a la lista de clientes
    if request.method == 'GET' and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return redirect('tienda:listar_clientes')
    # si el metodo es POST, guardamos el cliente en la base de datos
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tienda:listar_clientes')
        else:
            # Si hay errores, devolver status 400 para que el JS detecte el error
            return render(request, 'clientes/crear_cliente.html', {'form': form}, status=400)
    # si el metodo no es POST, mostramos el formulario para crear un nuevo cliente
    else:
        form = ClienteForm()
    return render(request, 'clientes/crear_cliente.html', {'form': form})

# funcion para crear un nuevo producto
def crear_producto(request):
    # si es get y no ajax, redirigir a la lista, ya que el formulario de creación de producto se muestra dentro de un modal en la lista de productos
    if request.method == 'GET' and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return redirect('tienda:listar_productos')
    # si el metodo es POST, guardamos el producto en la base de datos
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tienda:listar_productos')
        else:
            # Si hay errores, devolver status 400 para que el JS detecte el error
            return render(request, 'productos/crear_producto.html', {'form': form}, status=400)
    # si el metodo no es POST, mostramos el formulario para crear un nuevo producto
    else:
        form = ProductoForm()
    return render(request, 'productos/crear_producto.html', {'form': form})

# funcion para crear un nuevo pedido
def crear_pedido(request):
    # Si es GET y NO es AJAX, redirigir a la lista
    if request.method == 'GET' and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return redirect('tienda:listar_pedidos')
    
    # creamos la funcion de crear pedido pero le añadimos form set de detalles pedidos para que se asocie
    DetallePedidoFormSet = inlineformset_factory(Pedido, DetallePedido, form=DetallePedidoForm, extra=1, can_delete=False)
    # si el metodo es POST, guardamos el pedido y los detalles de pedido en la base de datos
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        formset = DetallePedidoFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            # Validar que al menos hay un detalle de pedido
            if not formset.has_changed():
                form.add_error(None, "Debes agregar al menos un producto al pedido.")
                return render(request, 'pedidos/crear_pedido.html', {'form': form, 'formset': formset}, status=400)
            
            pedido = form.save()
            detalles_pedido = formset.save(commit=False)
            for detalle in detalles_pedido:
                detalle.pedido = pedido
                # Recalcular precio en el servidor (seguridad)
                detalle.precio_pedido = detalle.producto.precio * detalle.cantidad
                detalle.save()
            return redirect('tienda:listar_pedidos')
        else:
            # Si hay errores, devolver status 400 para que el JS detecte el error
            return render(request, 'pedidos/crear_pedido.html', {'form': form, 'formset': formset}, status=400)
    # si el metodo no es POST, mostramos el formulario para crear un nuevo pedido con el form set de detalles pedidos
    else:
        form = PedidoForm()
        formset = DetallePedidoFormSet()
    return render(request, 'pedidos/crear_pedido.html', {'form': form, 'formset': formset})

# ver detalles de un cliente
def ver_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    return render(request, 'clientes/ver_cliente.html', {'cliente': cliente})

# ver detalles de un producto
def ver_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'productos/ver_producto.html', {'producto': producto})

# ver detalles de un pedido y se le asocia el detalle de pedido con el pedido para mostrarlo en la vista
def ver_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    detalles_pedido = DetallePedido.objects.filter(pedido=pedido)
    # datos de detalles_pedido como  precio formateados
    detalles_pedido = [
        {
            'producto': detalle.producto,
            'cantidad': detalle.cantidad,
            # el precio unitario viene del precio del producto
            'precio_unitario': f"{detalle.producto.precio:,.2f}",
            'precio_pedido': f"{detalle.precio_pedido:,.2f}",
        }
        for detalle in detalles_pedido
    ]
    return render(request, 'pedidos/ver_pedido.html', {'pedido': pedido, 'detalles_pedido': detalles_pedido})

# ver detalles de un detalle de pedido


# actualizar un cliente
def actualizar_cliente(request, pk):
    # obtenemos el cliente de la base de datos, si no existe, mostramos un error 404
    cliente = get_object_or_404(Cliente, pk=pk)
    # si el metodo es POST, actualizamos el cliente en la base de datos
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect('tienda:listar_clientes')
        else:
            # Si hay errores, devolver status 400 para que el JS detecte el error
            return render(request, 'clientes/actualizar_cliente.html', {'form': form}, status=400)
    # si el metodo no es POST, mostramos el formulario para actualizar el cliente con instancia del cliente a actualizar
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'clientes/actualizar_cliente.html', {'form': form})

# actualizar un producto
def actualizar_producto(request, pk):
    # obtenemos el producto de la base de datos, si no existe, mostramos un error 404
    producto = get_object_or_404(Producto, pk=pk)
    # si el metodo es POST, actualizamos el producto en la base de datos
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('tienda:listar_productos')
        else:
            # Si hay errores, devolver status 400 para que el JS detecte el error
            return render(request, 'productos/actualizar_producto.html', {'form': form}, status=400)
    # si el metodo no es POST, mostramos el formulario para actualizar el producto con instancia del producto a actualizar
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'productos/actualizar_producto.html', {'form': form})

# actualizar un pedido
def actualizar_pedido(request, pk):
    # obtenemos el pedido de la base de datos, si no existe, mostramos un error 404
    pedido = get_object_or_404(Pedido, pk=pk)
    # si el metodo es POST, actualizamos el pedido en la base de datos
    if request.method == 'POST':
        form = PedidoForm(request.POST, instance=pedido)
        if form.is_valid():
            form.save()
            return redirect('tienda:listar_pedidos')
        else:
            # Si hay errores, devolver status 400 para que el JS detecte el error
            return render(request, 'pedidos/actualizar_pedido.html', {'form': form}, status=400)
    # si el metodo no es POST, mostramos el formulario para actualizar el pedido con instancia del pedido a actualizar
    else:
        form = PedidoForm(instance=pedido)
    return render(request, 'pedidos/actualizar_pedido.html', {'form': form})

# actualizar un detalle de pedido
def actualizar_detalle_pedido(request, pk):
    # obtenemos el detalle de pedido de la base de datos, si no existe, mostramos un error 404
    detalle_pedido = get_object_or_404(DetallePedido, pk=pk)
    # si el metodo es POST, actualizamos el detalle de pedido en la base de datos
    if request.method == 'POST':
        form = DetallePedidoForm(request.POST, instance=detalle_pedido)
        if form.is_valid():
            form.save()
            return redirect('tienda:listar_detalles_pedidos')
        else:
            # Si hay errores, devolver status 400 para que el JS detecte el error
            return render(request, 'detalles_pedidos/actualizar_detalle_pedido.html', {'form': form}, status=400)
    # si el metodo no es POST, mostramos el formulario para actualizar el detalle de pedido con instancia del detalle de pedido a actualizar
    else:
        form = DetallePedidoForm(instance=detalle_pedido)
    return render(request, 'detalles_pedidos/actualizar_detalle_pedido.html', {'form': form})

# eliminar un cliente
def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if cliente.pedido_set.exists():
        # Si el cliente tiene pedidos asociados, no se puede eliminar (redirijomos a un sweet alert para mostrar el mensaje de error)
        # con un HttpResponse para mostrar el mensaje de error, esto es generico
        # return HttpResponse("No se puede eliminar el cliente debido a que tiene pedidos asociados.")
        return JsonResponse({  # utilizamos JsonResponse para enviar una respuesta JSON al frontend y mostrar el mensaje de error en un sweet alert
            'success': False,
            'message': "No se puede eliminar el cliente debido a que tiene pedidos asociados."
        })
    elif request.method == 'POST':
        cliente.delete()
        return JsonResponse({
            'success': True,
            'message': "Cliente eliminado exitosamente."
        })
        #return redirect('tienda:listar_clientes')
    #return render(request, 'clientes/eliminar_cliente.html', {'cliente': cliente})
    return JsonResponse({
        'success': False,
        'message': "Método no permitido."
    })

# eliminar un producto
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if producto.detallepedido_set.exists():
        # Si el producto tiene detalles de pedido asociados, no se puede eliminar
        return JsonResponse({
            'success': False,
            'message': "No se puede eliminar el producto debido a que tiene detalles de pedido asociados."
        })
    elif request.method == 'POST':
        producto.delete()
        return JsonResponse({
            'success': True,
            'message': "Producto eliminado exitosamente."
        })
    return JsonResponse({
        'success': False,
        'message': "Método no permitido."
    })

# eliminar un pedido
def eliminar_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    if pedido.detallepedido_set.exists():
        # Si el pedido tiene detalles asociados, no se puede eliminar
        return JsonResponse({
            'success': False,
            'message': "No se puede eliminar el pedido debido a que tiene detalles de pedido asociados."
        })
    elif request.method == 'POST':
        pedido.delete()
        return JsonResponse({
            'success': True,
            'message': "Pedido eliminado exitosamente."
        })
    return JsonResponse({
        'success': False,
        'message': "Método no permitido."
    })

# eliminar un detalle de pedido
def eliminar_detalle_pedido(request, pk):
    detalle_pedido = get_object_or_404(DetallePedido, pk=pk)
    detalle_pedido.delete()
    return redirect('tienda:listar_detalles_pedidos')

# reportes en excel de pedido por id  y sus detalles, incluido el cliente

def reporte_pedidos_excel(request):
    pedido_id = request.GET.get('pedido_id')
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    detalles_pedido = DetallePedido.objects.filter(pedido=pedido)

    # Crear un libro de Excel y una hoja
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Pedido_{pedido.id}"

    # Definir estilos
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_alignment = Alignment(horizontal="right", vertical="center")

    # Escribir los encabezados
    headers = ['ID Pedido', 'Cliente', 'Fecha', 'Estado', 'Producto', 'Cantidad', 'Subtotal']
    sheet.append(headers)

    # Formatear encabezados
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Calcular total
    total_pedido = sum(detalle.precio_pedido for detalle in detalles_pedido)
    
    # Escribir los datos del pedido y sus detalles
    for detalle in detalles_pedido:
        row = [
            pedido.id,
            pedido.cliente.nombre,
            pedido.fecha.strftime("%Y-%m-%d %H:%M"),
            pedido.estado,
            detalle.producto.nombre,
            detalle.cantidad,
            detalle.precio_pedido,
        ]
        sheet.append(row)

    # Formatear datos
    for row in sheet.iter_rows(min_row=2, max_row=len(detalles_pedido) + 1, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Formatear números como moneda solo en la columna de precio
            if cell.column == 7:
                cell.number_format = '$#,##0.00'
                cell.alignment = right_alignment
            if cell.column in [1, 6]:
                cell.alignment = center_alignment

    # Agregar fila de total
    total_row_num = len(detalles_pedido) + 2
    row = ['', '', '', '', 'TOTAL PEDIDO', '', total_pedido]
    sheet.append(row)

    # Formatear fila de total
    for cell in sheet[total_row_num]:
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if cell.column == 7:
            cell.number_format = '$#,##0.00'
            cell.alignment = right_alignment
        else:
            cell.alignment = center_alignment

    # Ajustar ancho de columnas
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 25
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 15

    # Guardar el libro de Excel en un objeto BytesIO
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Devolver el archivo Excel como respuesta HTTP
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Pedido_{pedido.id}.xlsx'
    return response

def factura_pedidos_detallepedidos_pdf(request):
    pedido_id = request.GET.get('pedido_id')
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    detalles_pedido = DetallePedido.objects.filter(pedido=pedido)

    # Crear un objeto BytesIO para almacenar el PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Estilos personalizados
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#2E75B6'),
        spaceAfter=10,
        alignment=1  # Centro
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#2E75B6'),
        spaceAfter=6,
        spaceBefore=10
    )
    
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=10,
        alignment=0
    )
    
    # Encabezado
    elements.append(Paragraph("FACTURA DE PEDIDO", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del pedido
    info_data = [
        [Paragraph("<b>Número de Pedido:</b>", normal_style), Paragraph(f"{pedido.id}", normal_style)],
        [Paragraph("<b>Cliente:</b>", normal_style), Paragraph(f"{pedido.cliente.nombre}", normal_style)],
        [Paragraph("<b>Teléfono:</b>", normal_style), Paragraph(f"{pedido.cliente.telefono}", normal_style)],
        [Paragraph("<b>Correo:</b>", normal_style), Paragraph(f"{pedido.cliente.correo}", normal_style)],
        [Paragraph("<b>Fecha:</b>", normal_style), Paragraph(f"{pedido.fecha.strftime('%d/%m/%Y %H:%M')}", normal_style)],
        [Paragraph("<b>Estado:</b>", normal_style), Paragraph(f"{pedido.estado}", normal_style)],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 2.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8EEF7')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de detalles
    elements.append(Paragraph("DETALLES DEL PEDIDO", heading_style))
    
    detail_data = [[
        Paragraph('<b>Producto</b>', normal_style),
        Paragraph('<b>Cantidad</b>', normal_style),
        Paragraph('<b>Precio Unitario</b>', normal_style),
        Paragraph('<b>Subtotal</b>', normal_style)
    ]]
    
    total_pedido = 0
    for detalle in detalles_pedido:
        subtotal = detalle.precio_pedido
        total_pedido += subtotal
        detail_data.append([
            Paragraph(detalle.producto.nombre, normal_style),
            Paragraph(str(detalle.cantidad), normal_style),
            Paragraph(f"${detalle.producto.precio:,.2f}", normal_style),
            Paragraph(f"${subtotal:,.2f}", normal_style)
        ])
    
    # Agregar fila de total
    detail_data.append([
        Paragraph('', normal_style),
        Paragraph('', normal_style),
        Paragraph('<b>TOTAL:</b>', normal_style),
        Paragraph(f'<b>${total_pedido:,.2f}</b>', normal_style)
    ])
    
    detail_table = Table(detail_data, colWidths=[3*inch, 1.2*inch, 1.3*inch, 1.3*inch])
    detail_table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Datos
        ('ALIGN', (0, 1), (0, -2), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -2), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -2), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 6),
        ('TOPPADDING', (0, 1), (-1, -2), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        
        # Total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFC000')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ('TOPPADDING', (0, -1), (-1, -1), 8),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
    ]))
    
    elements.append(detail_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Pie de página
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=1
    )
    elements.append(Paragraph("Gracias por su compra • Documento generado automáticamente", footer_style))
    
    # Construir el PDF
    doc.build(elements)
    
    # Devolver el PDF como respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Factura_Pedido_{pedido.id}.pdf'
    return response
    
# funcion para traer datos a un dashboard
def dashboard(request):
    total_clientes = Cliente.objects.count()
    total_pedidos = Pedido.objects.count()
    # se obtiene desde detalle el producto, pero como se envia el id, con __nombre se puede mostrar el nombre del producto en la vista y annotate es para contar la cantidad total vendida de cada producto, luego se ordena de mayor a menor y se obtiene el primero para mostrar el producto mas vendido (el - es para ordenar de mayor a menor)
    producto_mas_vendido = DetallePedido.objects.values('producto__nombre').annotate(total_vendido=Sum('cantidad')).order_by('-total_vendido').first()
    producto_menos_vendido = DetallePedido.objects.values('producto__nombre').annotate(total_vendido=Sum('cantidad')).order_by('total_vendido').first()
    
    # top 3 clientes con mas pedidos sin contar los cancelados
    # el ~Q es de django para negar la consulta, es decir, contar los pedidos que no tengan estado cancelado
    top_clientes = Cliente.objects.annotate(num_pedidos=Count('pedido', filter=~Q(pedido__estado='Cancelado'))).order_by('-num_pedidos')[:3]
    
    # pedidos por estado
    pedidos_por_estado = Pedido.objects.values('estado').annotate(count=Count('id'))
    ventas_por_mes = Pedido.objects.filter(estado='Entregado').annotate(mes=TruncMonth('fecha')).values('mes').annotate(total_ventas=Sum('detallepedido__precio_pedido')).order_by('mes')
    
    # total de ventas solo de estado "Entregado"
    total_ventas = Pedido.objects.filter(estado='Entregado').aggregate(total=Sum('detallepedido__precio_pedido'))['total'] or 0
    
    context = {
        'total_clientes': total_clientes,
        'total_pedidos': total_pedidos,
        'top_clientes': top_clientes,
        'pedidos_por_estado': pedidos_por_estado,
        'total_ventas': f"{total_ventas:,.0f}".replace('.', ','),
        'producto_mas_vendido': producto_mas_vendido,
        'producto_menos_vendido': producto_menos_vendido,
        'ventas_por_mes': 
        [
            {
                'mes': venta['mes'].strftime('%B %Y') if venta['mes'] else 'Sin fecha',
                'total_ventas': f"{venta['total_ventas'] or 0:,.0f}".replace('.', ',')
            }
            for venta in ventas_por_mes
        ]
    }
    return render(request, 'dashboard/dashboard.html', context)
    
    
    
    
